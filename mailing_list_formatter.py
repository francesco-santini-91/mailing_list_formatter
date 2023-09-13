# Nome: 		mailing_list_formatter.py
# Autore: 		Francesco Santini
# Versione: 	1.0
# Data: 		06/09/2023
# Descrizione: 	Script per la formattazione di un CSV destinato all'importazione di mailing list
# Utilizzo: 	Eseguire lo script lanciando il comando python <nome_file.xlsx> colonna1=dataType1 colonna2 = dataType2... Indicando il nome della colonna presente sul foglio Excel ed il tipo di dato, 
#				utilizzando "email" per le colonne contenenti gli indirizzi email, e "text" per tutte le altre.
# Esempio:		python test.xlsx ragione_sociale=text cognome=text nome=text email1=email

import os
import sys
import openpyxl
import csv
import re

filename = sys.argv[1]
headers = []
dataTypes = []
inserted = []
success = 0
errors = 0
duplicate = 0

for i in range(2, len(sys.argv)):
	arg = sys.argv[i].split("=")
	if (len(arg) == 2):
		if (arg[1] == "text" or arg[1] == "email"):
			headers.append(arg[0])
			dataTypes.append(arg[1])
		else:
			sys.exit("Errore! Tipo di dato non valido: " + arg[1])
	else:
		sys.exit("Errore! Intestazioni non corrette.")

if not (os.path.isfile(filename)):
	sys.exit("Errore! Impossibile aprire il file " + filename)

excel = openpyxl.load_workbook(filename)
sheet = excel.active

exported_file = open(filename.split(".")[0] + ".csv", "w", newline = "", encoding = "UTF8")
errors_file =  open(filename.split(".")[0] + "_errors.csv", "w", newline = "", encoding = "UTF8")
exported_writer = csv.writer(exported_file, delimiter = ";")
errors_writer = csv.writer(errors_file, delimiter = ";")
exported_writer.writerow(headers)

for r in range(2, sheet.max_row + 1):
	tmp = [None] * len(headers)
	err = [None] * int(len(headers) + 1)
	for index, column in enumerate(headers):
		if (dataTypes[index] == "text"):
			tmp[index] = sheet.cell(row = r, column = index + 1).value
			err[index] = tmp[index]
		if (dataTypes[index] == "email"):
			if (sheet.cell(row = r, column = index + 1).value):
				regex = re.compile(r'\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,7}\b')
				addresses = re.findall(regex, sheet.cell(row = r, column = index + 1).value)
				if (len(addresses)):
					tmp[index] = addresses
				else:
					err[index] = addresses
					err[len(headers)] = "Indirizzo email non valido."
	emails = tmp[dataTypes.index("email")]
	if (hasattr(emails, "__iter__")):
		for email in emails:
			occourrences = [match for match in inserted if email.lower() in match]
			if not occourrences:
				row = [None] * len(headers)
				for index, column in enumerate(headers):
					if (dataTypes[index] == "text"):
						row[index] = tmp[index].strip().capitalize() if tmp[index] else ""
					if (dataTypes[index] == "email"):
						row[index] = email.lower()
				exported_writer.writerow(row)
				inserted.append(email.lower())
				success = success + 1
			else:
				duplicate = duplicate + 1
	else:
		err[len(headers)] = "Indirizzo email non presente."
		errors_writer.writerow(err)
		errors = errors + 1

exported_file.close()
errors_file.close()

print("Indirizzi email esportati: " + str(success))
print("Indirizzi email non validi: " + str(errors))
print("Indirizzi email duplicati: " + str(duplicate))
